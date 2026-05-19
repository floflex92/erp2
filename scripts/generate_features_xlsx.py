from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


CSV_PATH = Path(r"C:\Users\Florent\erp2\erp2\docs\modele-fonctionnalites-page.csv")
XLSX_PATH = Path(r"C:\Users\Florent\erp2\erp2\docs\modele-fonctionnalites-page.xlsx")


def main() -> None:
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))

    if not rows:
        raise ValueError("CSV is empty")

    wb = Workbook()
    ws = wb.active
    ws.title = "Fonctionnalites"

    for row in rows:
        ws.append(row)

    header = rows[0]
    max_row = ws.max_row
    max_col = ws.max_column

    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = "A2"

    widths: dict[int, int] = {}
    for r in rows:
        for i, value in enumerate(r, start=1):
            widths[i] = max(widths.get(i, 0), len(value))

    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 12), 48)

    col_idx = {name: idx + 1 for idx, name in enumerate(header)}

    def add_list_validation(column_name: str, values: list[str]) -> None:
        idx = col_idx.get(column_name)
        if not idx:
            return

        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=True,
            showDropDown=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(idx)}2:{get_column_letter(idx)}500")

    add_list_validation("Statut", ["Developpe", "En cours de developpement", "Features"])
    add_list_validation("Priorite", ["Haute", "Moyenne", "Basse"])
    add_list_validation("Publier_sur_site", ["TRUE", "FALSE"])

    wb.save(XLSX_PATH)


if __name__ == "__main__":
    main()
